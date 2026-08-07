from __future__ import annotations

import io
import re
import sqlite3
import unicodedata
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from .db import log_activity, now_iso

READINESS_STAGES = [
    "Business zadání",
    "Příprava promptu",
    "Externí dependance",
    "Test (lokální)",
    "E2E test",
    "Pilot",
    "Business roll-out",
]

SKIP_SHEETS = {"schedule", "BOARD launch to production"}

SHEET_ALIASES = {
    "BondingOptika": "Bonding a optika",
    "Bonding wHW": "Bonding + HW",
    "mojeO2": "MOA",
    "Unity": "Unity",
    "expSlevy": "Expirace slev",
    "O2SPOLUcold_leads": "O2 Spolu",
    "Alenka": "Alenka",
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _text(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", _norm(value)).strip("-")
    return slug or "project"


def _date(value: Any) -> str | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m."):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == "%d.%m.":
                parsed = parsed.replace(year=datetime.now().year)
            return parsed.date().isoformat()
        except ValueError:
            pass
    return text[:40]


def _health_from_readiness(statuses: list[str]) -> str:
    normalized = [_norm(s) for s in statuses if s]
    if any(x in {"issue", "blocked", "blocker"} or "issue" in x for x in normalized):
        return "Red"
    if any(x in {"in progress", "to do", "todo", "not started"} for x in normalized):
        return "Amber"
    if normalized and all(x in {"done", "completed"} for x in normalized):
        return "Green"
    return "Amber"


def _project_type(sheet_name: str) -> str:
    n = _norm(sheet_name)
    if any(key in n for key in ["monitoring", "bigquery", "integrace", "procurement", "deduplikace", "nakupni cesta"]):
        return "Enabler"
    return "Campaign"


def _find_header_row(ws) -> tuple[int | None, dict[str, int]]:
    aliases = {
        "area": {"oblast", "area"},
        "title": {"task", "ukol"},
        "detail": {"detail", "popis"},
        "status": {"status", "termin"},
        "blocks": {"blokuje go live", "blokuje golive", "blocker"},
        "owner": {"owner", "owner ", "vlastnik"},
        "date": {"date", "datum", "termin"},
        "next": {"next action detail", "next action", "next steps", "dalsi krok"},
        "comment": {"comment", "komentar"},
    }
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        mapping: dict[str, int] = {}
        for col_idx in range(1, min(ws.max_column, 12) + 1):
            cell_norm = _norm(ws.cell(row_idx, col_idx).value)
            if not cell_norm:
                continue
            for key, values in aliases.items():
                if cell_norm in values and key not in mapping:
                    mapping[key] = col_idx
        if "title" in mapping or ("detail" in mapping and "owner" in mapping):
            return row_idx, mapping
    return None, {}


def _read_meta(ws) -> tuple[str, str, str, str]:
    title = ""
    business_owner = ""
    spoc = ""
    summary = ""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 6), values_only=True):
        for value in row[:4]:
            text = _text(value)
            norm = _norm(text)
            if not text:
                continue
            if not title and not norm.startswith(("business owner", "spoc", "spok", "aidcc")):
                title = text
            if norm.startswith("business owner"):
                business_owner = text.split(":", 1)[1].strip() if ":" in text else text
            if norm.startswith(("spoc", "spok", "spoc aidcc")):
                spoc = text.split(":", 1)[1].strip() if ":" in text else text
    if title:
        summary = title
    return title, business_owner, spoc, summary


def _upsert_project(
    conn: sqlite3.Connection,
    name: str,
    source_ref: str,
    business_owner: str = "",
    spoc: str = "",
    project_type: str = "Campaign",
    summary: str = "",
    health: str = "Amber",
) -> int:
    row = conn.execute("SELECT id FROM projects WHERE source_ref = ? OR name = ?", (source_ref, name)).fetchone()
    now = now_iso()
    if row:
        project_id = int(row[0])
        conn.execute(
            """
            UPDATE projects SET name=?, business_owner=?, aidcc_spoc=?, project_type=?, summary=?, health=?, updated_at=?
            WHERE id=?
            """,
            (name, business_owner, spoc, project_type, summary, health, now, project_id),
        )
        return project_id

    base = _slug(name)
    slug = base
    counter = 2
    while conn.execute("SELECT 1 FROM projects WHERE slug=?", (slug,)).fetchone():
        slug = f"{base}-{counter}"
        counter += 1
    cursor = conn.execute(
        """
        INSERT INTO projects(name, slug, project_type, business_owner, aidcc_spoc, status, health, summary, source_ref, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, 'Active', ?, ?, ?, ?, ?)
        """,
        (name, slug, project_type, business_owner, spoc, health, summary, source_ref, now, now),
    )
    return int(cursor.lastrowid)


def import_workbook(conn: sqlite3.Connection, raw: bytes) -> dict[str, int]:
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
    stats = {"projects": 0, "actions": 0, "readiness": 0}

    # Demo records are useful only before the first real workbook import.
    conn.execute("DELETE FROM projects WHERE source_ref LIKE 'demo:%'")
    readiness_by_name: dict[str, list[tuple[str, str]]] = {}

    if "Status" in wb.sheetnames:
        ws = wb["Status"]
        headers = [_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            project_name = _text(ws.cell(r, 1).value)
            if not project_name:
                continue
            stages: list[tuple[str, str]] = []
            for c in range(2, min(ws.max_column, 8) + 1):
                stage = headers[c - 1] or (READINESS_STAGES[c - 2] if c - 2 < len(READINESS_STAGES) else f"Stage {c-1}")
                status = _text(ws.cell(r, c).value) or "Not started"
                stages.append((stage, status))
            readiness_by_name[_norm(project_name)] = stages

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS or sheet_name == "Status":
            continue
        ws = wb[sheet_name]
        title, business_owner, spoc, summary = _read_meta(ws)
        project_type = _project_type(sheet_name)
        canonical_name = SHEET_ALIASES.get(sheet_name)
        if canonical_name:
            display_name = canonical_name
        elif project_type == "Enabler":
            display_name = sheet_name
        else:
            display_name = title or sheet_name
            if len(display_name) > 70 or _norm(display_name) in {"oblast", "task", "detail"}:
                display_name = sheet_name

        readiness_key = canonical_name or display_name
        matched_stages = readiness_by_name.get(_norm(readiness_key)) or readiness_by_name.get(_norm(sheet_name)) or readiness_by_name.get(_norm(display_name)) or []
        health = _health_from_readiness([status for _, status in matched_stages]) if matched_stages else "Amber"
        project_id = _upsert_project(
            conn,
            display_name,
            f"xlsx:{sheet_name}",
            business_owner,
            spoc,
            project_type,
            summary if summary != display_name else "",
            health,
        )
        stats["projects"] += 1

        conn.execute("DELETE FROM readiness WHERE project_id=?", (project_id,))
        for idx, (stage, status) in enumerate(matched_stages):
            conn.execute(
                "INSERT INTO readiness(project_id, stage, status, sort_order) VALUES (?, ?, ?, ?)",
                (project_id, stage, status, idx),
            )
            stats["readiness"] += 1

        conn.execute("DELETE FROM actions WHERE project_id=? AND source_ref LIKE 'xlsx:%'", (project_id,))
        header_row, columns = _find_header_row(ws)
        if not header_row:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            def val(key: str) -> Any:
                col = columns.get(key)
                return ws.cell(r, col).value if col else None

            area = _text(val("area"))
            title_value = _text(val("title"))
            detail = _text(val("detail"))
            title_value = title_value or area or (detail[:120] if detail else "")
            if not title_value:
                continue
            status = _text(val("status")) or "To Do"
            owner = _text(val("owner"))
            blocks_raw = _norm(val("blocks"))
            blocks = 1 if blocks_raw in {"ano", "yes", "true", "1", "blocker", "ad blocker"} else 0
            due = _date(val("date"))
            next_action = _text(val("next"))
            comment = _text(val("comment"))
            now = now_iso()
            conn.execute(
                """
                INSERT INTO actions(project_id, area, title, detail, status, blocks_go_live, owner, due_date, next_action, comment, source_ref, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    area,
                    title_value,
                    detail,
                    status,
                    blocks,
                    owner,
                    due,
                    next_action,
                    comment,
                    f"xlsx:{sheet_name}:{r}",
                    now,
                    now,
                ),
            )
            stats["actions"] += 1

        log_activity(conn, project_id, "project", project_id, "Imported from Excel", sheet_name)

    return stats
